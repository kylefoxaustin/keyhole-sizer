"""
KPI breakdown per pipeline — backend support for the Streamlit "download KPI
spreadsheet" feature.

Emits one flat dict per (pipeline, hardware, resolution, LLM-state) with
columns that mirror Kyle's requested KPI table:

  model | ingest_type | ingest_ms | yolo_type | yolo_ms |
  seg_type | seg_ms | llm_type | llm_ms | total_fps

This module is pure-data + pure-function. The Streamlit UI side (app.py)
composes rows, serializes to CSV/XLSX, and wires `st.download_button`.
"""
from __future__ import annotations

import io
from typing import Any, Optional

from sizer.npu_model import (
    PIPELINES, Hardware, project_vision, project_llm,
    CLIP_FP8_EDGE_MS_NPU_MID,
)


# ───────────────────────── Constants ─────────────────────────

# FFmpeg decode + letterbox resize to 640×640 on an edge ARM core.
# Measured on 5090 host (i9-14900KF, OpenCV 4.11, single-thread): 0.17 ms
# at 720p / 0.32 ms at 1080p / 0.33 ms at 4K. Edge ARM (Cortex-A55 class,
# ~10× slower single-thread): ~2–3 ms per frame, near-flat across source
# resolutions because the output (640×640) dominates the bilinear cost.
# See `keyhole-sizer/sizer/measured.py` + the deck slide #43 preprocessing
# disclosure for provenance.
INGEST_MS_EDGE_720P = 2.0
INGEST_MS_EDGE_1080P = 2.5
INGEST_MS_EDGE_4K = 3.0

INGEST_MS_EDGE_BY_RES = {
    "720p": INGEST_MS_EDGE_720P,
    "1080p": INGEST_MS_EDGE_1080P,
    "4K": INGEST_MS_EDGE_4K,
}

INGEST_TYPE_LABEL = "FFmpeg + letterbox"


# ───────────────────────── Stage attribution table ─────────────────────────
#
# For each pipeline key we name the YOLO-side model (detector or combined
# detect+seg) and the segmentation-side model (separate mask model, or
# CLIP tagger for hybrid/TRT pipelines, or None for YOLO-seg-only).
#
# The latency split is handled by the row-builder using two strategies:
#
# 1. **Composed pipelines** (YOLO + CLIP): `project_vision()` already
#    returns `yolo_ms` and `clip_ms` — we use those directly.
# 2. **SAM-lineage pipelines**: project_vision returns only a single
#    per_stream_ms. We attribute a CONSTANT edge-ms chunk to the Yolo11x
#    detector (~30 ms/frame at NPU Mid, flat across source resolutions
#    since YOLO-seg always sees 640²), and the rest to the SAM stage.
#    That constant is stored here per-pipeline so a calibrated Yolo11x
#    profile can be swapped in later without touching the builder.
# 3. **One-model pipelines** (YOLOE-26): the single YOLO forward does
#    both detect + open-vocab, so yolo_ms = pipeline edge_ms, seg_ms = 0.
# 4. **YOLO-seg-only** (yolo_only_fp8, yolov8n_only_fp8, INT8 variants):
#    same as #3 — yolo_ms = edge_ms, no separate seg stage.

# Yolo11x detector cost at edge — used by SAM-lineage pipelines.
# Derived from the mask-model bake-off's yolo_detection stage and scaled
# via the NPU-Mid BW ratio. See `bakeoff_sam_variants.py` for the raw 5090
# measurement that this extrapolates from.
YOLO11X_DETECT_MS_EDGE = 30.0


PIPELINE_STAGES: dict[str, dict[str, Any]] = {
    # ─── SAM 3 lineage — YOLO11x detects, SAM model segments ──────────
    "sam3_bf16": {
        "category":   "sam",
        "yolo_type":  "Yolo11x",
        "seg_type":   "SAM 3 BF16",
    },
    "essmall_fp8": {
        "category":   "sam",
        "yolo_type":  "Yolo11x",
        "seg_type":   "EfficientSAM-Small FP8",
    },
    "efficientsam3_es_ev_s_bf16": {
        "category":   "sam",
        "yolo_type":  "Yolo11x",
        "seg_type":   "EfficientSAM3 ES-EV-S BF16",
    },
    "efficientsam3p1_es_ev_s_bf16": {
        "category":   "sam",
        "yolo_type":  "(text prompt)",
        "seg_type":   "EfficientSAM3.1 ES-EV-S BF16",
    },

    # ─── One-model open-vocab — single forward does both ──────────────
    "yoloe26_s_pf_fp16": {
        "category":   "one_model",
        "yolo_type":  "YOLOE-26S-PF FP16",
        "seg_type":   None,
    },
    "yoloe26_s_pf_trt_fp8": {
        "category":   "one_model",
        "yolo_type":  "YOLOE-26S-PF TRT FP8",
        "seg_type":   None,
    },

    # ─── Shipping lineage — YOLO-seg + CLIP enrich ─────────────────────
    "hybrid_v2_bf16": {
        "category":   "composed",
        "yolo_type":  "Yolo11s-seg BF16",
        "seg_type":   "CLIP BF16 (every frame)",
    },
    "hybrid_v2_torchao_fp8": {
        "category":   "composed",
        "yolo_type":  "Yolo11s-seg BF16",
        "seg_type":   "CLIP FP8 torchao (every frame)",
    },
    "trt_fp8_every_frame": {
        "category":   "composed",
        "yolo_type":  "Yolo11s-seg FP8 TRT",
        "seg_type":   "CLIP FP8 TRT (every frame)",
    },
    "trt_fp8_1hz_clip": {
        "category":   "composed",
        "yolo_type":  "Yolo11s-seg FP8 TRT",
        "seg_type":   "CLIP FP8 TRT (1 Hz)",
    },
    "yolo_only_fp8": {
        "category":   "yolo_only",
        "yolo_type":  "Yolo11s-seg FP8 TRT",
        "seg_type":   None,
    },

    # ─── yolov8n nano — YOLO-seg does both; CLIP optional ──────────────
    "yolov8n_trt_fp8_every_frame": {
        "category":   "composed",
        "yolo_type":  "Yolov8n-seg FP8 TRT",
        "seg_type":   "CLIP FP8 TRT (every frame)",
    },
    "yolov8n_trt_fp8_1hz_clip": {
        "category":   "composed",
        "yolo_type":  "Yolov8n-seg FP8 TRT",
        "seg_type":   "CLIP FP8 TRT (1 Hz)",
    },
    "yolov8n_only_fp8": {
        "category":   "yolo_only",
        "yolo_type":  "Yolov8n-seg FP8 TRT",
        "seg_type":   None,
    },

    # ─── INT8 vendor-comparison — yolo-only, calibration-tagged ────────
    "yolo11s_trt_int8": {
        "category":   "yolo_only",
        "yolo_type":  "Yolo11s-seg INT8 TRT (20-frame PTQ)",
        "seg_type":   None,
    },
    "yolov8n_trt_int8_coco128": {
        "category":   "yolo_only",
        "yolo_type":  "Yolov8n-seg INT8 TRT (coco128-seg PTQ)",
        "seg_type":   None,
    },
    "resnet50v1_int8_224": {
        # Pure image classification — no detector head. Treated as
        # "one_model" (single-forward pipeline) for KPI categorization.
        "category":   "one_model",
        "yolo_type":  "ResNet-50v1 INT8 TRT (ImageNet 224×224)",
        "seg_type":   None,
    },
    # 4-bit-weight variants added 2026-05-14 to unlock the spec's
    # {mid,high}_int8.{resnet50_w4,yolov8n_w4} anchor cells per
    # [docs] 20:56 — same KPI categories as their 8-bit siblings;
    # only the weight quantization differs.
    "yolov8n_trt_int4_coco128": {
        "category":   "yolo_only",
        "yolo_type":  "Yolov8n-seg INT4-w TRT (coco128-seg PTQ)",
        "seg_type":   None,
    },
    "resnet50v1_int4_224": {
        "category":   "one_model",
        "yolo_type":  "ResNet-50v1 INT4-w TRT (ImageNet 224×224)",
        "seg_type":   None,
    },

    # ─── ViT alternatives (Kyle 2026-04-25 what-if) ───────────────────
    # Camera-stream candidates: single-model object detectors. Treated
    # as "one_model" (full pipeline time = single forward). Agentic
    # candidates: text-prompted detection, also one_model in shape.
    "rtdetr_l_pytorch_fp16": {
        "category":   "one_model",
        "yolo_type":  "RT-DETR-L FP16 (ViT)",
        "seg_type":   None,
    },
    "detr_resnet50_pytorch_fp16": {
        "category":   "one_model",
        "yolo_type":  "DETR ResNet-50 FP16 (ViT)",
        "seg_type":   None,
    },
    "owlv2_base_pytorch_fp16": {
        "category":   "one_model",
        "yolo_type":  "OWLv2-base FP16 (open-vocab ViT)",
        "seg_type":   None,
    },
    "grounding_dino_tiny_pytorch_fp32": {
        "category":   "one_model",
        "yolo_type":  "Grounding DINO Tiny FP32 (text-grounded ViT)",
        "seg_type":   None,
    },
}


def _stage_split_ms(pipeline_key: str, vision_result: dict,
                     pipeline_total_ms: float) -> tuple[float, float]:
    """Return (yolo_ms, seg_ms) for this pipeline at the active HW/resolution.

    Drives the "yolo_ms" + "seg_ms" columns in the KPI row. See the module
    docstring for the attribution strategy per category.
    """
    category = PIPELINE_STAGES.get(pipeline_key, {}).get("category", "unknown")

    if category == "composed":
        # project_vision returns a breakdown for these — use it directly.
        y = vision_result.get("yolo_ms", 0.0)
        c = vision_result.get("clip_ms", 0.0)
        return y, c

    if category == "yolo_only" or category == "one_model":
        # Single-stage pipelines — the full pipeline time is the YOLO stage.
        return pipeline_total_ms, 0.0

    if category == "sam":
        # Attribute a flat Yolo11x chunk to detection; remainder to seg.
        y = YOLO11X_DETECT_MS_EDGE
        s = max(0.0, pipeline_total_ms - y)
        return y, s

    # Unknown — safest is to dump everything into yolo and log 0 for seg
    return pipeline_total_ms, 0.0


def pipeline_kpi_row(
    pipeline_key: str,
    hw: Hardware,
    resolution: str,
    llm_enabled: bool = False,
    llm_quant: str = "Q4_K_M",
    llm_workload: str = "plain_chat",
    compiler_quality_vs_trt: float = 1.0,
) -> dict[str, Any]:
    """Emit one KPI row for a single (pipeline × hw × resolution × LLM) point.

    Returns a flat dict suitable for direct serialization to a CSV/XLSX row.
    Column keys are stable and match Kyle's requested spreadsheet shape.

    `total_fps` is VISION-only (no LLM duty-cycle deduction). If llm_enabled,
    the LLM columns populate with per-query timings and quant/workload labels,
    but total_fps does not apply a haircut — the user can duty-cycle in the
    sizer's interactive UI. Keeps rows stable and interpretable for
    deployment-math purposes.
    """
    if pipeline_key not in PIPELINES:
        raise KeyError(f"Unknown pipeline_key: {pipeline_key}")
    if pipeline_key not in PIPELINE_STAGES:
        raise KeyError(
            f"pipeline_key={pipeline_key!r} is registered in PIPELINES but not "
            f"in PIPELINE_STAGES. Add a stage attribution entry to "
            f"sizer/kpi_breakdown.py before it can appear in the KPI row."
        )
    pipeline = PIPELINES[pipeline_key]
    stage_meta = PIPELINE_STAGES[pipeline_key]

    # Vision projection on this hw/res (single stream)
    vision = project_vision(pipeline, hw, resolution, n_streams=1,
                             compiler_quality_vs_trt=compiler_quality_vs_trt)
    pipeline_total_ms = vision["per_stream_ms"]
    yolo_ms, seg_ms = _stage_split_ms(pipeline_key, vision, pipeline_total_ms)

    # Ingest cost (edge ARM letterbox)
    ingest_ms = INGEST_MS_EDGE_BY_RES[resolution]

    # LLM cost (per short-answer query, if enabled)
    if llm_enabled:
        llm = project_llm(hw, llm_quant, workload=llm_workload)
        # project_llm returns *_sec keys (seconds), not ms. Convert.
        short_sec = llm.get("short_answer_sec", 0.0)
        llm_ms = short_sec * 1000.0
        llm_type = f"Qwen3-30B-A3B {llm_quant}"
    else:
        llm_ms = 0.0
        llm_type = "off"

    # Total pipeline vision FPS (ingest + yolo + seg; LLM NOT in the
    # per-frame budget — it duty-cycles, see docstring).
    vision_ms = ingest_ms + yolo_ms + seg_ms
    total_fps = (1000.0 / vision_ms) if vision_ms > 0 else 0.0

    return {
        "model":          pipeline.label,
        "pipeline_key":   pipeline_key,
        "hw":             hw.name,
        "resolution":     resolution,
        "ingest_type":    INGEST_TYPE_LABEL,
        "ingest_ms":      round(ingest_ms, 2),
        "yolo_type":      stage_meta["yolo_type"],
        "yolo_ms":        round(yolo_ms, 2),
        "seg_type":       stage_meta["seg_type"] if stage_meta["seg_type"] else "n/a",
        "seg_ms":         round(seg_ms, 2) if stage_meta["seg_type"] else None,
        "llm_type":       llm_type,
        "llm_ms":         round(llm_ms, 2),
        "total_fps":      round(total_fps, 2),
    }


def all_pipeline_kpi_rows(
    hw: Hardware,
    resolution: str = "720p",
    llm_enabled: bool = False,
    llm_quant: str = "Q4_K_M",
    llm_workload: str = "plain_chat",
    compiler_quality_vs_trt: float = 1.0,
) -> list[dict[str, Any]]:
    """Emit KPI rows for every pipeline in PIPELINES at a fixed (hw, res,
    LLM-state) point. Ordered by the stage-attribution categories (SAM
    lineage → one-model → composed → yolo_only → INT8), which matches
    the deck's narrative order.
    """
    category_order = {"sam": 0, "one_model": 1, "composed": 2, "yolo_only": 3, "unknown": 99}
    keys_sorted = sorted(
        (k for k in PIPELINES if k in PIPELINE_STAGES),
        key=lambda k: (category_order.get(PIPELINE_STAGES[k]["category"], 99), k),
    )
    return [
        pipeline_kpi_row(k, hw, resolution,
                         llm_enabled=llm_enabled,
                         llm_quant=llm_quant,
                         llm_workload=llm_workload,
                         compiler_quality_vs_trt=compiler_quality_vs_trt)
        for k in keys_sorted
    ]


# ─────── XLSX export: formatted spreadsheet ─────────────────────────
#
# CSV can't carry bold/alignment, so when Kyle wants a human-readable
# downloadable spreadsheet we emit XLSX via openpyxl. Formatting rules:
#   • Row 1 (headers): Title Case, bold
#   • Column A: right-aligned
#   • All other columns: center-aligned
#   • Numeric columns keep numeric type (sortable in Excel/Calc)
#
# openpyxl is pinned in requirements.txt (added alongside this module).

# Column key → display header mapping. Handles acronyms so we don't get
# clown-case output like "Hw" or "Ms" — those stay uppercase/lowercase
# as readers expect them.
_HEADER_DISPLAY: dict[str, str] = {
    "model":         "Model",
    "pipeline_key":  "Pipeline Key",
    "hw":            "HW",
    "resolution":    "Resolution",
    "ingest_type":   "Ingest Type",
    "ingest_ms":     "Ingest ms",
    "yolo_type":     "YOLO Type",
    "yolo_ms":       "YOLO ms",
    "seg_type":      "Seg Type",
    "seg_ms":        "Seg ms",
    "llm_type":      "LLM Type",
    "llm_ms":        "LLM ms",
    "total_fps":     "Total FPS",
}


def kpi_rows_to_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Render KPI rows to an XLSX byte buffer with Kyle's requested formatting:

      • Row 1 is Title-Cased + bold.
      • Column A is right-aligned.
      • All other columns are center-aligned.
      • Column widths auto-sized to content.

    Returns raw bytes ready for `st.download_button`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    if not rows:
        # Empty workbook with just headers from the mapping
        rows = [{k: "" for k in _HEADER_DISPLAY}]

    # Preserve column order from the first row's dict (Python dicts are
    # insertion-ordered since 3.7). If _HEADER_DISPLAY has a mapping for
    # the key, use that; otherwise fall back to the raw key.
    columns = list(rows[0].keys())
    headers = [_HEADER_DISPLAY.get(c, c.replace("_", " ").title()) for c in columns]

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    colA_align = Alignment(horizontal="right", vertical="center")
    other_align = Alignment(horizontal="center", vertical="center")

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_align

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(columns, start=1):
            val = row.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = colA_align if col_idx == 1 else other_align

    # Auto-size column widths based on content (cap to a sensible max)
    for col_idx, key in enumerate(columns, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        values = [headers[col_idx - 1]] + [
            str(r.get(key, "")) if r.get(key) is not None else "" for r in rows
        ]
        max_len = max((len(v) for v in values), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def all_pipeline_kpi_xlsx(
    hw: Hardware,
    resolution: str = "720p",
    llm_enabled: bool = False,
    llm_quant: str = "Q4_K_M",
    llm_workload: str = "plain_chat",
    compiler_quality_vs_trt: float = 1.0,
) -> bytes:
    """Convenience: full-sweep KPI rows + XLSX bytes in one call. Drop-in
    replacement for the CSV path in app.py — same signature."""
    rows = all_pipeline_kpi_rows(hw, resolution,
                                  llm_enabled=llm_enabled,
                                  llm_quant=llm_quant,
                                  llm_workload=llm_workload,
                                  compiler_quality_vs_trt=compiler_quality_vs_trt)
    return kpi_rows_to_xlsx(rows)


# ─────── Startup assertion: stage attribution contract ───────────────
# If a pipeline is added to PIPELINES without an accompanying
# PIPELINE_STAGES entry, the KPI spreadsheet would silently drop it.
# Mirrors the `PIPELINES ⊆ ⋃(tracks)` invariant [sizer] added in
# app.py commit 8c696a2 — fail loud at import so stale attribution is
# caught at deploy time, not mid-spreadsheet-generation.
_orphaned = set(PIPELINES) - set(PIPELINE_STAGES)
if _orphaned:
    raise RuntimeError(
        f"sizer/kpi_breakdown.py: pipelines registered in PIPELINES but "
        f"missing stage attribution: {sorted(_orphaned)}. Add each to "
        f"PIPELINE_STAGES with (yolo_type, seg_type, category)."
    )
_unknown = set(PIPELINE_STAGES) - set(PIPELINES)
if _unknown:
    raise RuntimeError(
        f"sizer/kpi_breakdown.py: PIPELINE_STAGES references pipelines "
        f"not in PIPELINES: {sorted(_unknown)}. Remove or fix the key."
    )
